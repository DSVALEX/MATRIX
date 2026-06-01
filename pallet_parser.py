"""
pallet_parser.py — parse the DHL multi-country pallet rate card.

Source format (one sheet, "Sheet1"):
    Country | Zip | Country + Zip | 0,1 - 100 kg | 100,1 - 200 kg | ... | FTL

Each row is one destination zone (Country + Zip outcode); each weight column is a
band whose UPPER bound is encoded in the header. Rates are already in EUR and
already factored (the source is "DHL_pricing_with_factor..."), so no currency
conversion is applied here — the per-carrier `factor` in PALLET_DEFAULTS divides
the value if a manager wants to re-factor.

Public API
----------
is_pallet_file(path)              -> bool
parse_pallet_rate_card(path)      -> dict   {ISO2: {zone: {band_upper_kg: rate}}}
country_pallet_data(pallets, iso2)-> dict   {zone: {band_upper_kg: rate}}  (or {})
available_pallet_countries(p)     -> set
"""

import re
import logging

import openpyxl

log = logging.getLogger(__name__)

_BAND_RE = re.compile(r'-\s*(\d+)\s*kg', re.IGNORECASE)


def _band_upper(header):
    """'700,1 - 800 kg' -> 800 ; 'FTL' -> 'FTL' ; else None."""
    if header is None:
        return None
    h = str(header).strip()
    if h.upper() == 'FTL':
        return 'FTL'
    m = _BAND_RE.search(h)
    return int(m.group(1)) if m else None


def is_pallet_file(path):
    """True if this looks like the DHL pallet grid (Country/Zip + weight-band cols)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return False
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    if not header:
        return False
    norm = [str(h).strip().lower() if h is not None else '' for h in header[:3]]
    has_keys = ('country' in norm and any('zip' in n for n in norm))
    has_bands = sum(1 for h in header if _band_upper(h) not in (None,)) >= 5
    return has_keys and has_bands


def parse_pallet_rate_card(path):
    """Parse the whole pallet card once.

    Returns {ISO2: {zone: {band_upper_kg: rate, ..., 'FTL': rate}}}.
    `zone` is the Zip/outcode string exactly as in the source (e.g. 'B', '10').
    Single-zone countries (LU, MT, TR) keep their single zone key.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = list(rows[0])

    # Map each data column index -> band upper bound (skip the 3 key columns)
    band_cols = {}
    for ci, h in enumerate(header):
        if ci < 3:
            continue
        up = _band_upper(h)
        if up is not None:
            band_cols[ci] = up

    out = {}
    for r in rows[1:]:
        country = r[0]
        zone = r[1]
        if country is None or zone is None:
            continue
        iso2 = str(country).strip().upper()
        zone = str(zone).strip()
        bands = {}
        for ci, up in band_cols.items():
            v = r[ci] if ci < len(r) else None
            if isinstance(v, (int, float)):
                bands[up] = float(v)
        if bands:
            out.setdefault(iso2, {})[zone] = bands

    log.info('pallet card: %d countries, %d zones total',
             len(out), sum(len(z) for z in out.values()))
    return out


def country_pallet_data(pallets, iso2):
    """Return {zone: {band_upper_kg: rate}} for one country (empty if absent)."""
    return pallets.get(iso2.upper(), {})


def available_pallet_countries(pallets):
    """All ISO2 codes with pallet rate data."""
    return set(pallets.keys())
