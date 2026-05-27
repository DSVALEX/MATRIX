"""
pipeline.py — Rate Matrix Builder core module.

All logic lives here. The Streamlit app (app.py) imports and calls run_pipeline().
Key design decision: every function that reads CARRIER_DEFAULTS or VARIABLES_LAYOUT
accepts an optional override so the app can customise rates per session without
touching module-level globals (important for concurrent Streamlit users).
"""

import re
import logging
import shutil
from copy import deepcopy
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

log = logging.getLogger(__name__)


# ==============================================================================
# 1. CONFIGURATION
# ==============================================================================

CARRIER_DEFAULTS = {
    'UPDE': {
        'label': 'UPS DE',
        'services': ['STANDARD', 'EXPRESS SAVER 7R9W62', 'EXPRESS SAVER'],
        'has_postcode': True,
        'volume_divisor': 167,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B1',
        'maut_pct': 0.0,
        'maut_variables_ref': None,
        'linehaul_per_parcel': 1.31,
    },
    'UPSNL': {
        'label': 'UPS NL',
        'services': ['EXPRESS SAVER'],
        'has_postcode': True,
        'volume_divisor': 167,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B4',
        'maut_pct': 0.0,
        'maut_variables_ref': None,
        'linehaul_per_parcel': 0.0,
    },
    'DPD': {
        'label': 'DPD',
        'services': ['PARCEL'],
        'has_postcode': False,
        'volume_divisor': 250,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B3',
        'maut_pct': 0.05,
        'maut_variables_ref': 'B8',
        'linehaul_per_parcel': 0.0,
    },
    'DHL-ROS': {
        'label': 'DHL',
        'services': ['STANDARD'],
        'has_postcode': False,
        'volume_divisor': 250,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B2',
        'maut_pct': 0.06,
        'maut_variables_ref': 'B9',
        'linehaul_per_parcel': 0.0,
    },
    'POSTNORD': {
        'label': 'PostNord',
        'services': ['STANDARD'],
        'has_postcode': False,
        'volume_divisor': 250,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B5',
        'maut_pct': 0.0,
        'maut_variables_ref': None,
        'linehaul_per_parcel': 0.0,
    },
    'UPSGB': {
        'label': 'UPS GB',
        'services': ['STANDARD', 'EXPRESS SAVER'],
        'has_postcode': False,
        'volume_divisor': 167,
        'fuel_pct': 0.27,
        'fuel_variables_ref': 'B6',
        'maut_pct': 0.0,
        'maut_variables_ref': None,
        'linehaul_per_parcel': 3.9,
    },
}

VARIABLES_LAYOUT = [
    ('FUEL UPSDE',    0.27),   # B1
    ('FUEL DHL',      0.27),   # B2
    ('FUEL DPD',      0.27),   # B3
    ('FUEL UPSNL',    0.27),   # B4
    ('FUEL POSTNORD', 0.27),   # B5
    ('FUEL UPSGB',    0.27),   # B6
    (None, None),              # B7  blank spacer
    ('MAUT DPD',      0.05),   # B8
    ('MAUT DHL',      0.06),   # B9
]

_BASE_CARRIERS   = ['UPDE', 'DPD', 'DHL-ROS', 'UPSNL']
_SCANDI_CARRIERS = ['UPDE', 'DPD', 'DHL-ROS', 'UPSNL', 'POSTNORD']
_SCANDI_ISO      = {'SE', 'DK', 'NO', 'FI'}


def _default_country_cfg(iso2):
    return {
        'iso2':                  iso2,
        'site_id':               'NLMOE01',
        'client_id':             'NLFENDER',
        'max_parcel_count':      10,
        'max_each_weight_kg':    31.5,
        'each_weight_grid':      sorted(set(list(range(1, 32)) + [31.5])),
        'carriers':              _SCANDI_CARRIERS if iso2 in _SCANDI_ISO else _BASE_CARRIERS,
        'postcode_prefix_range': (0, 99),
    }


COUNTRY_CONFIG = {iso: _default_country_cfg(iso) for iso in [
    'DE', 'FR', 'IT', 'ES', 'NL',
    'BE', 'IE', 'PT', 'LU',
    'AT', 'CH', 'PL', 'CZ', 'SK', 'HU', 'SI',
    'SE', 'DK', 'NO', 'FI',
    'GR', 'HR', 'BG', 'RO', 'SM',
    'EE', 'LV', 'LT',
]}
# DE only has three carriers (no UPSNL)
COUNTRY_CONFIG['DE']['carriers'] = ['UPDE', 'DPD', 'DHL-ROS']

# GB — UK domestic via UPSGB, plus NL-origin export carriers that quote GB
COUNTRY_CONFIG['GB'] = _default_country_cfg('GB')
COUNTRY_CONFIG['GB']['carriers'] = ['UPSGB', 'UPDE', 'DPD', 'DHL-ROS', 'UPSNL']


# ==============================================================================
# 2. ROBUST TEXT / SHEET HELPERS
# ==============================================================================

_JUNK = re.compile(r'[\s\-_/\\.,;:()\[\]]+')


def _norm(v):
    """Normalise any value: lowercase, collapse all punctuation/whitespace to ' '."""
    if v is None:
        return ''
    return _JUNK.sub(' ', str(v).lower()).strip()


def _cell_match(cell_value, *needles):
    """True if the normalised cell value contains/equals ANY normalised needle."""
    cv = _norm(cell_value)
    if not cv:
        return False
    for raw in needles:
        n = _norm(raw)
        if n and (n in cv or cv in n):
            return True
    return False


def _parse_float(v):
    """Parse float robustly: handles comma-decimals, currency symbols, None."""
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        raise ValueError('None')
    s = re.sub(r'[€$£\s]', '', str(v).strip())
    if ',' in s and '.' in s:
        if s.index(',') > s.index('.'):        # '1.234,56' European
            s = s.replace('.', '').replace(',', '.')
        else:                                   # '1,234.56' Anglo
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    return float(s)


def _find_sheet(wb, *name_hints):
    """Return first sheet whose name fuzzy-matches any hint; None if not found."""
    for hint in name_hints:
        if hint in wb.sheetnames:
            return wb[hint]
    for hint in name_hints:
        hn = _norm(hint)
        for sname in wb.sheetnames:
            if hn and (_norm(sname) == hn or hn in _norm(sname)):
                return wb[sname]
    return None


def _scan_anchor(ws, *anchor_texts, max_row=None, max_col=None):
    """First cell that fuzzy-matches any anchor text; None if not found."""
    mr = max_row or ws.max_row
    mc = max_col or ws.max_column
    for r in range(1, mr + 1):
        for c in range(1, mc + 1):
            v = ws.cell(r, c).value
            if v is not None and _cell_match(v, *anchor_texts):
                return (r, c)
    return None


# keep original name as alias for backward compat
_scan_for_anchor = _scan_anchor

_FROM_WORDS = {'from', 'van', 'von', 'de', 'fra', 'vanaf', 'weight from', 'kg from'}
_TO_WORDS   = {'to', 'tot', 'bis', 'a', 'til', 'tot en met', 'weight to', 'kg to'}


def _find_from_to(ws, anchor_row, anchor_col, max_rows=10, col_slack=5):
    """Find From/To header row near anchor. Multi-language, wider search window."""
    col_lo = max(1, anchor_col - col_slack)
    col_hi = min(anchor_col + col_slack, ws.max_column)
    row_lo = max(1, anchor_row - 1)
    row_hi = min(anchor_row + max_rows, ws.max_row)

    for r in range(row_lo, row_hi + 1):
        for c in range(col_lo, col_hi + 1):
            cv = _norm(ws.cell(r, c).value)
            if cv not in _FROM_WORDS:
                continue
            for to_offset in (1, 2):
                if _norm(ws.cell(r, c + to_offset).value) in _TO_WORDS:
                    return r, c, c + to_offset
    return None, None, None


_find_from_to_header = _find_from_to   # alias

_OVER_RE = re.compile(
    r'\b(over|meer|plus|above|mehr|oltre|mas|vidare|sup[ée]rieur)\b|\+\s*$'
)


def _extract_tiers(ws, header_row, from_col, to_col, rate_col, max_rows=200):
    """Extract weight-band tiers; tolerates blank spacers, comma-decimals,
    all 'over/meer/plus' end markers."""
    tiers = []
    last_from = -1
    blanks = 0

    for r in range(header_row + 1, min(ws.max_row, header_row + max_rows) + 1):
        f_raw = ws.cell(r, from_col).value
        t_raw = ws.cell(r, to_col).value
        rate_raw = ws.cell(r, rate_col).value

        if f_raw is None and rate_raw is None:
            blanks += 1
            if blanks > 2:
                break
            continue
        blanks = 0

        try:
            f_val = _parse_float(f_raw)
        except (TypeError, ValueError):
            break

        if tiers and f_val < last_from and f_val <= 1:
            break
        last_from = f_val

        try:
            rate_val = _parse_float(rate_raw)
        except (TypeError, ValueError):
            continue

        t_str = _norm(t_raw) if t_raw is not None else ''
        if t_raw is None or _OVER_RE.search(t_str):
            tiers.append({'from': f_val, 'to': float('inf'),
                          'rate': rate_val, 'per_kg': True})
            break
        try:
            t_val = _parse_float(t_raw)
        except (TypeError, ValueError):
            break

        tiers.append({'from': f_val, 'to': t_val, 'rate': rate_val, 'per_kg': False})

    return tiers


_extract_tier_table = _extract_tiers   # alias


def _extract_rates_by_zone(ws, hrow, from_col, to_col):
    by_zone = {}
    for c in range(to_col + 1, ws.max_column + 1):
        v = ws.cell(hrow, c).value
        if v is None:
            if by_zone:
                break
            continue
        if isinstance(v, (int, float)):
            tiers = _extract_tiers(ws, hrow, from_col, to_col, c)
            if tiers:
                by_zone[int(v)] = tiers
            continue
        if isinstance(v, str):
            vs = v.strip()
            if _norm(vs) in ('from', 'to', 'payweight', 'van', 'tot'):
                break
            if re.fullmatch(r'[A-Z]{2,3}', vs):
                tiers = _extract_tiers(ws, hrow, from_col, to_col, c)
                if tiers:
                    by_zone[vs] = tiers
                continue
            if by_zone:
                break
    return by_zone


# ==============================================================================
# 3. RATE-CARD PARSER
# ==============================================================================

def _parse_upsde_zones(ws):
    anchor = _scan_anchor(ws, 'Zones UPSDE', 'ZONES UPSDE', 'zones ups de')
    if not anchor:
        return []
    ar, ac = anchor
    hrow, from_col, to_col = _find_from_to(ws, ar, ac, max_rows=4, col_slack=4)
    if not hrow:
        return []
    country_col = from_col - 1
    service_cols = {}
    for label_row in (hrow, hrow - 1):
        if label_row < 1:
            continue
        for c in range(to_col + 1, ws.max_column + 1):
            v = ws.cell(label_row, c).value
            if not isinstance(v, str):
                continue
            vl = _norm(v)
            if 'standard single' in vl and 'STDS' not in service_cols:
                service_cols['STDS'] = c
            elif 'standard multi' in vl and 'STDM' not in service_cols:
                service_cols['STDM'] = c
            elif 'express saver' in vl and 'EXPRESS_SAVER' not in service_cols:
                service_cols['EXPRESS_SAVER'] = c
        if service_cols:
            break

    zones = []
    for r in range(hrow + 1, ws.max_row + 1):
        country = ws.cell(r, country_col).value
        pc_from = ws.cell(r, from_col).value
        pc_to   = ws.cell(r, to_col).value
        if country is None and pc_from is None and pc_to is None:
            break
        if isinstance(pc_from, str) and pc_from.strip().upper() == 'ALL':
            pc_from_int, pc_to_int = 0, 99999
        else:
            try:
                pc_from_int = int(str(pc_from))
                pc_to_int   = int(str(pc_to)) if pc_to is not None else pc_from_int
            except (TypeError, ValueError):
                continue
        entry = {'country': country, 'pc_from': pc_from_int, 'pc_to': pc_to_int}
        for svc, col in service_cols.items():
            v = ws.cell(r, col).value
            if isinstance(v, (int, float)):
                entry[svc] = int(v)
            elif isinstance(v, str) and v.strip():
                entry[svc] = v.strip()
        zones.append(entry)
    return zones


# PostNord service anchors
_PN_SHEET_NAMES   = ['POSTNORD', 'POST NORD', 'PostNord', 'Post Nord', 'PN']
_PN_SVC_ANCHORS   = {
    'STANDARD': ['parcel postnord', 'postnord standard', 'postnord parcel',
                 'rate per parcel postnord', 'standard postnord', 'postnord'],
    'EXPRESS':  ['postnord express', 'express postnord'],
    'ECONOMY':  ['postnord economy', 'economy postnord'],
}


def _parse_postnord_sheet(ws):
    """
    Four-strategy PostNord parser.

    Strategy 0 — Flat rate per service code (e.g. SE: B2B|18|15P → 12.2)
        Looks for a 'Servicelevel' header; collects service-code rows where
        the right column is numeric.  Returns 'flat_rates': {name: rate}.

    Strategy 1 — Anchored service tier tables (From/To weight bands)
    Strategy 2 — Full-sheet From/To scan
    Strategy 3 — Numeric region detection
    """
    result = {}

    # ── Strategy 0: flat rate per service code ────────────────────────────────
    anchor = _scan_anchor(ws, 'servicelevel', 'service level', 'service code')
    if anchor:
        ar, ac = anchor
        for svc_col in (ac, ac + 1):
            flat_rates = {}
            for r in range(ar + 1, ws.max_row + 1):
                svc_raw = ws.cell(r, svc_col).value
                if svc_raw is None:
                    continue
                svc_str = str(svc_raw).strip()
                # Skip pure country-code rows like 'SE', 'DK'
                if re.fullmatch(r'[A-Z]{2,3}', svc_str):
                    continue
                # Skip header-like strings
                if _norm(svc_str) in ('to country', 'country', 'to', 'rate',
                                      'service', 'servicelevel'):
                    continue
                for rate_offset in (1, 2):
                    try:
                        rate = _parse_float(ws.cell(r, svc_col + rate_offset).value)
                        svc_name = svc_str.split('|')[0].strip().upper()
                        if svc_name:
                            flat_rates[svc_name] = rate
                        break
                    except (TypeError, ValueError):
                        continue
            if flat_rates:
                result['flat_rates'] = flat_rates
                log.info('    PostNord strategy 0: flat rates %s', flat_rates)
                return result

    # ── Strategy 1: anchored service tier tables ──────────────────────────────
    for svc, anchors in _PN_SVC_ANCHORS.items():
        anchor = _scan_anchor(ws, *anchors)
        if not anchor:
            continue
        ar, ac = anchor
        hrow, fc, tc = _find_from_to(ws, ar, ac, max_rows=12, col_slack=6)
        if not hrow:
            continue
        by_zone = _extract_rates_by_zone(ws, hrow, fc, tc)
        if by_zone:
            result[f'{svc}_by_zone'] = by_zone
            result[svc] = next(iter(by_zone.values()))
        else:
            tiers = _extract_tiers(ws, hrow, fc, tc, tc + 1)
            if tiers:
                result[svc] = tiers
        if svc in result:
            break

    if result:
        return result

    # ── Strategy 2: full-sheet From/To scan ──────────────────────────────────
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(r, c).value) not in _FROM_WORDS:
                continue
            for off in (1, 2):
                if _norm(ws.cell(r, c + off).value) not in _TO_WORDS:
                    continue
                tiers = _extract_tiers(ws, r, c, c + off, c + off + 1)
                if len(tiers) >= 2:
                    result['STANDARD'] = tiers
                    return result

    # ── Strategy 3: numeric region detection ─────────────────────────────────
    for sc in range(1, max(1, ws.max_column - 2) + 1):
        for sr in range(1, ws.max_row + 1):
            run = 0
            for rr in range(sr, min(sr + 30, ws.max_row) + 1):
                try:
                    _parse_float(ws.cell(rr, sc).value)
                    _parse_float(ws.cell(rr, sc + 1).value)
                    _parse_float(ws.cell(rr, sc + 2).value)
                    run += 1
                except (TypeError, ValueError):
                    break
            if run >= 3:
                tiers = _extract_tiers(ws, sr - 1, sc, sc + 1, sc + 2)
                if len(tiers) >= 2:
                    result['STANDARD'] = tiers
                    return result

    log.warning('PostNord: all strategies failed on sheet "%s"', ws.title)
    return result


def parse_rate_cards(excel_path):
    """Parse all carrier rate tables from the uploaded Excel. Fuzzy sheet matching."""
    wb  = openpyxl.load_workbook(excel_path, data_only=True)
    out = {}

    # ── UPSDE ────────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, 'UPSDE', 'UPS DE', 'UPS-DE', 'UPDE')
    if ws:
        upde = {'zones': _parse_upsde_zones(ws)}
        for label, key in [('PARCEL - UPS - STDS', 'STDS'),
                            ('PARCEL - UPS - STDM', 'STDM')]:
            anchor = _scan_anchor(ws, label)
            if anchor:
                ar, ac = anchor
                hrow, fc, tc = _find_from_to(ws, ar, ac)
                if hrow:
                    upde[f'{key}_by_zone'] = _extract_rates_by_zone(ws, hrow, fc, tc)
                    upde[key] = _extract_tiers(ws, hrow, fc, tc, tc + 1)

        anchor = _scan_anchor(ws, 'EXPSAVER UPSDE 7R9W62', 'EXPSAVER 7R9W62')
        if anchor:
            ar, ac = anchor
            for r in range(ar + 1, min(ar + 6, ws.max_row) + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and re.fullmatch(r'[A-Z]{2}', v.strip()):
                        try:
                            upde['EXPSAVER_7R9W62'] = _parse_float(ws.cell(r, c + 1).value)
                            break
                        except (TypeError, ValueError):
                            pass
                if 'EXPSAVER_7R9W62' in upde:
                    break

        anchor = _scan_anchor(ws, 'UPS DE - LINEHAUL', 'UPSDE Linehaul', 'linehaul ups')
        if anchor:
            ar, ac = anchor
            for r in range(ar + 1, min(ar + 6, ws.max_row) + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and any(
                        kw in v.upper() for kw in ('DEUTSCHLAND', 'GERMANY', 'DE')
                    ):
                        try:
                            upde['LINEHAUL'] = _parse_float(ws.cell(r, c + 1).value)
                            break
                        except (TypeError, ValueError):
                            pass
                if 'LINEHAUL' in upde:
                    break

        anchor = _scan_anchor(ws, 'EXPRESS SAVER UPSDE', 'PARCEL - EXPRESS SAVER UPSDE')
        if anchor:
            ar, ac = anchor
            hrow, fc, tc = _find_from_to(ws, ar, ac)
            if hrow:
                upde['EXPRESS_SAVER_by_zone'] = _extract_rates_by_zone(ws, hrow, fc, tc)
                upde['EXPRESS_SAVER'] = _extract_tiers(ws, hrow, fc, tc, tc + 1)

        out['UPDE'] = upde

    # ── UPSNL ────────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, 'UPSNL', 'UPS NL', 'UPS-NL')
    if ws:
        upsnl = {'zones': [], 'rates_by_zone': {}}
        anchor = _scan_anchor(ws, 'ZONES UPSNL', 'Zones UPSNL')
        if anchor:
            ar, ac = anchor
            hrow, fc, tc = _find_from_to(ws, ar, ac, max_rows=5, col_slack=4)
            if hrow:
                country_col = fc - 1
                zone_col = None
                for c in range(tc + 1, ws.max_column + 1):
                    if _cell_match(ws.cell(hrow, c).value, 'express saver'):
                        zone_col = c
                        break
                if zone_col is None:
                    for c in range(tc + 1, ws.max_column + 1):
                        if isinstance(ws.cell(hrow + 1, c).value, (int, float)):
                            zone_col = c
                            break
                if zone_col:
                    for r in range(hrow + 1, ws.max_row + 1):
                        pc_from  = ws.cell(r, fc).value
                        pc_to    = ws.cell(r, tc).value
                        zone_raw = ws.cell(r, zone_col).value
                        if pc_from is None and pc_to is None and zone_raw is None:
                            break
                        if isinstance(pc_from, str) and pc_from.strip().upper() == 'ALL':
                            pc_from_int, pc_to_int = 0, 99999
                        else:
                            try:
                                pc_from_int = int(str(pc_from))
                                pc_to_int   = int(str(pc_to))
                            except (TypeError, ValueError):
                                continue
                        try:
                            upsnl['zones'].append({
                                'country':  ws.cell(r, country_col).value,
                                'pc_from':  pc_from_int,
                                'pc_to':    pc_to_int,
                                'zone':     int(str(zone_raw)),
                            })
                        except (TypeError, ValueError):
                            continue

        anchor = _scan_anchor(ws, 'PARCEL - EXPRESS SAVER UPSNL', 'EXPRESS SAVER UPSNL',
                              'Rates UPSNL express saver', 'Rates UPSNL')
        if anchor:
            ar, ac = anchor
            hrow, fc, tc = _find_from_to(ws, ar, ac, max_rows=4, col_slack=3)
            if hrow:
                rbz = _extract_rates_by_zone(ws, hrow, fc, tc)
                upsnl['rates_by_zone'] = {k: v for k, v in rbz.items()
                                          if isinstance(k, int)}
        out['UPSNL'] = upsnl

    # ── DHL ──────────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, 'DHL', 'DHL-ROS', 'DHL ROS')
    if ws:
        anchor = _scan_anchor(ws, 'PARCEL - DHL', 'Rate per parcel DHL', 'dhl standard')
        if anchor:
            ar, ac = anchor
            hrow, fc, tc = _find_from_to(ws, ar, ac)
            if hrow:
                tiers = _extract_tiers(ws, hrow, fc, tc, tc + 1)
                if tiers:
                    out['DHL-ROS'] = {'STANDARD': tiers}

    # ── DPD ──────────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, 'DPD')
    if ws:
        anchor = _scan_anchor(ws, 'PARCEL - DPD', 'Rate per parcel DPD', 'dpd parcel')
        if anchor:
            ar, _ = anchor
            label_map = {'groot': 'groot', 'klein': 'klein', 'big': 'groot',
                         'small': 'klein', 'large': 'groot', 'heavy': 'groot'}
            rates = {}
            for r in range(ar + 1, min(ar + 6, ws.max_row) + 1):
                row_labels = {}
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str):
                        k = label_map.get(v.strip().lower())
                        if k:
                            row_labels[c] = k
                if row_labels:
                    for col, norm_label in row_labels.items():
                        try:
                            rates[norm_label] = _parse_float(ws.cell(r - 1, col).value)
                        except (TypeError, ValueError):
                            pass
                    break
            if rates:
                out['DPD'] = rates

    # ── POSTNORD ─────────────────────────────────────────────────────────────
    ws = _find_sheet(wb, *_PN_SHEET_NAMES)
    if ws:
        data = _parse_postnord_sheet(ws)
        if data:
            out['POSTNORD'] = data
        else:
            log.warning('POSTNORD sheet found but no rates parsed')

    return out


# ==============================================================================
# 4. TIER UTILITIES
# ==============================================================================

def lookup_tier_rate(tiers, weight):
    if weight <= 0:
        return None
    for t in tiers:
        if t['from'] < weight <= t['to']:
            return t['rate'] * weight if t['per_kg'] else t['rate']
    if tiers and weight == tiers[0]['from']:
        return tiers[0]['rate']
    return None


def collapse_same_rate_tiers(tiers, weight_cap=None):
    if not tiers:
        return []
    bands, cur_rate, cur_to, cur_pk = [], tiers[0]['rate'], tiers[0]['to'], tiers[0].get('per_kg', False)
    for t in tiers[1:]:
        if t['rate'] == cur_rate and not cur_pk:
            cur_to = t['to']
        else:
            bands.append((cur_to, cur_rate, cur_pk))
            cur_rate, cur_to, cur_pk = t['rate'], t['to'], t.get('per_kg', False)
    bands.append((cur_to, cur_rate, cur_pk))
    if weight_cap is not None:
        bands = [(min(to, weight_cap), rate, pk) for to, rate, pk in bands
                 if to == float('inf') or to <= weight_cap * 1.5]
        bands = [(weight_cap if to == float('inf') else to, rate, pk)
                 for to, rate, pk in bands]
        seen, out = set(), []
        for to, rate, pk in bands:
            if (to, rate) not in seen:
                seen.add((to, rate))
                out.append((to, rate, pk))
        bands = out
    return bands


# ==============================================================================
# 5. MATRIX BUILDERS
# ==============================================================================

def _upde_service_buckets(rate_data, service_key, country_cfg):
    """Return [(postcode_prefix_or_None, tiers), ...] for a UPDE service.

    Postcode-resistance rules:
    - No zones at all            → flat rate, no postcode column
    - Zones but no pc_from/pc_to → treat as single zone, no postcode column
    - All zones map to same rate → collapse to single entry, no postcode column
    - Multiple distinct zones    → one entry per postcode prefix in range
    - Prefix not covered by any  → falls back to the worst (highest-rate) zone
      so no order is ever left unmatched
    """
    zones   = rate_data.get('zones', [])
    by_zone = rate_data.get(f'{service_key}_by_zone', {})
    flat    = rate_data.get(service_key, [])
    pc_min, pc_max = country_cfg['postcode_prefix_range']

    if not zones:
        return [(None, flat)] if flat else []

    zone_ids = [z[service_key] for z in zones if service_key in z]
    if not zone_ids:
        return [(None, flat)] if flat else []
    unique_zones = set(zone_ids)

    def tiers_for(zid):
        if isinstance(zid, int):
            return by_zone.get(zid, [])
        return by_zone.get(zid) or flat

    # If zones carry no pc_from/pc_to (old-style single-zone entry), treat as flat
    has_pc_ranges = any('pc_from' in z and 'pc_to' in z for z in zones)
    if not has_pc_ranges or len(unique_zones) == 1:
        return [(None, tiers_for(next(iter(unique_zones))))]

    # Find the fallback: worst zone (highest rate at max weight)
    def _zone_max_rate(zid):
        t = tiers_for(zid)
        return max((b['rate'] for b in t), default=0)
    fallback_zid = max(unique_zones, key=_zone_max_rate)

    buckets = []
    for pc in range(pc_min, pc_max + 1):
        pc_full = pc * 1000
        zid = next((z[service_key] for z in zones
                    if z.get('pc_from', 0) <= pc_full <= z.get('pc_to', 99999)
                    and service_key in z), fallback_zid)
        t = tiers_for(zid)
        if t:
            buckets.append((pc, t))

    # If all prefixes resolved to the same zone, collapse to no-postcode
    if buckets and len({t_id for _, t_id in
                        [(pc, id(t)) for pc, t in buckets]}) == 1:
        return [(None, buckets[0][1])]

    return buckets


def _common(site, client, carrier, iso2):
    return {'SITE_ID': site, 'CLIENT_ID': client, 'CARRIER_ID': carrier,
            'COUNTRYISO2': iso2, 'POSTCODE': None, 'MIN_WEIGHT': None,
            'MIN_VOLUME': None, 'MIN_PARCEL': None,
            'USER_DEF_TYPE_4 (max 1,5m)': None, 'AWKWARD': None, 'RATE_EXTRA': 0}


def build_rows_upde(rate_data, country_cfg):
    rows   = []
    max_p  = country_cfg['max_parcel_count']
    max_ew = country_cfg['max_each_weight_kg']
    c0     = _common(country_cfg['site_id'], country_cfg['client_id'],
                     'UPDE', country_cfg['iso2'])
    seen_by_pc = {}

    for pc, tiers in _upde_service_buckets(rate_data, 'STDS', country_cfg):
        seen = seen_by_pc.setdefault(pc, set())
        for each_w, rate, per_kg in collapse_same_rate_tiers(tiers, max_ew):
            if each_w > max_ew or per_kg:
                continue
            for mp in range(1, max_p + 1):
                rb  = rate * mp
                key = (mp, each_w, rb)
                if key not in seen:
                    seen.add(key)
                    rows.append({**c0, 'POSTCODE': pc, 'SERVICE_LEVEL': 'STANDARD',
                                 'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                                 'RATE_BASE': round(rb, 4)})

    for pc, tiers in _upde_service_buckets(rate_data, 'STDM', country_cfg):
        seen = seen_by_pc.setdefault(pc, set())
        for combined_w, rate, per_kg in collapse_same_rate_tiers(tiers):
            if per_kg:
                continue
            for mp in range(1, max_p + 1):
                each_w = combined_w / mp
                if each_w not in country_cfg['each_weight_grid'] or each_w > max_ew:
                    continue
                key = (mp, each_w, rate)
                if key not in seen:
                    seen.add(key)
                    rows.append({**c0, 'POSTCODE': pc, 'SERVICE_LEVEL': 'STANDARD',
                                 'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                                 'RATE_BASE': round(rate, 4)})

    flat = rate_data.get('EXPSAVER_7R9W62')
    if flat is not None:
        for mp in range(1, max_p + 1):
            rows.append({**c0, 'SERVICE_LEVEL': 'EXPRESS SAVER 7R9W62',
                         'MAX_PARCEL': mp, 'EACH_WEIGHT': max_ew,
                         'RATE_BASE': round(flat * mp, 4)})

    for pc, tiers in _upde_service_buckets(rate_data, 'EXPRESS_SAVER', country_cfg):
        for each_w, rate, per_kg in collapse_same_rate_tiers(tiers, max_ew):
            if each_w > max_ew or per_kg:
                continue
            for mp in range(1, max_p + 1):
                rows.append({**c0, 'POSTCODE': pc, 'SERVICE_LEVEL': 'EXPRESS SAVER',
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                             'RATE_BASE': round(rate * mp, 4)})

    # ---- WorldEase (WEA): flat per-country rate (CH, NO) ----
    wea = rate_data.get('WEA')
    if wea is not None:
        for mp in range(1, max_p + 1):
            rows.append({**c0, 'SERVICE_LEVEL': 'WORLDEASE',
                         'MAX_PARCEL': mp, 'EACH_WEIGHT': max_ew,
                         'RATE_BASE': round(wea * mp, 4)})

    return rows


def build_rows_dhl(rate_data, country_cfg):
    rows  = []
    max_p = country_cfg['max_parcel_count']
    max_ew= country_cfg['max_each_weight_kg']
    c0    = _common(country_cfg['site_id'], country_cfg['client_id'],
                    'DHL-ROS', country_cfg['iso2'])

    # ---- BNL pricing: 1st parcel + each-additional, no weight tiers (BE/LU/NL) ----
    bnl = rate_data.get('bnl')
    if bnl:
        first = bnl.get('first')
        after = bnl.get('after', first)
        if first is not None:
            for mp in range(1, max_p + 1):
                total = first + (mp - 1) * after
                rows.append({**c0, 'SERVICE_LEVEL': 'STANDARD',
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': max_ew,
                             'RATE_BASE': round(total, 4)})
        return rows

    for each_w, rate, per_kg in collapse_same_rate_tiers(
            rate_data.get('STANDARD', []), max_ew):
        if each_w > max_ew or per_kg:
            continue
        for mp in range(1, max_p + 1):
            rows.append({**c0, 'SERVICE_LEVEL': 'STANDARD',
                         'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                         'RATE_BASE': round(rate * mp, 4)})
    return rows


def build_rows_dpd(rate_data, country_cfg):
    rows  = []
    max_p = country_cfg['max_parcel_count']
    c0    = _common(country_cfg['site_id'], country_cfg['client_id'],
                    'DPD', country_cfg['iso2'])
    for mp in range(1, max_p + 1):
        if 'klein' in rate_data:
            rows.append({**c0, 'SERVICE_LEVEL': 'PARCEL', 'MAX_PARCEL': mp,
                         'EACH_WEIGHT': 3.0,
                         'RATE_BASE': round(rate_data['klein'] * mp, 4)})
        if 'groot' in rate_data:
            rows.append({**c0, 'SERVICE_LEVEL': 'PARCEL', 'MAX_PARCEL': mp,
                         'EACH_WEIGHT': 31.5,
                         'RATE_BASE': round(rate_data['groot'] * mp, 4)})
    return rows


def build_rows_upsnl(rate_data, country_cfg):
    rows   = []
    max_p  = country_cfg['max_parcel_count']
    max_ew = country_cfg['max_each_weight_kg']
    pc_min, pc_max = country_cfg['postcode_prefix_range']
    c0     = _common(country_cfg['site_id'], country_cfg['client_id'],
                     'UPSNL', country_cfg['iso2'])

    zones = rate_data.get('zones', [])
    bands_by_zone = {z: collapse_same_rate_tiers(t, max_ew)
                     for z, t in rate_data.get('rates_by_zone', {}).items()}

    # No zone table at all → no postcode, use zone 1 (or first available)
    if not zones:
        fallback_zone = next(iter(bands_by_zone), None)
        if fallback_zone is None:
            return rows
        for each_w, rate, per_kg in bands_by_zone[fallback_zone]:
            if each_w > max_ew or per_kg:
                continue
            for mp in range(1, max_p + 1):
                rows.append({**c0, 'POSTCODE': None, 'SERVICE_LEVEL': 'EXPRESS SAVER',
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                             'RATE_BASE': round(rate * mp, 4)})
        return rows

    # Has pc_from/pc_to ranges → map each prefix to a zone
    has_pc_ranges = any('pc_from' in z and 'pc_to' in z for z in zones)

    if not has_pc_ranges:
        # Single-zone entry without postcode ranges → no postcode column
        zone = zones[0].get('zone')
        for each_w, rate, per_kg in bands_by_zone.get(zone, []):
            if each_w > max_ew or per_kg:
                continue
            for mp in range(1, max_p + 1):
                rows.append({**c0, 'POSTCODE': None, 'SERVICE_LEVEL': 'EXPRESS SAVER',
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                             'RATE_BASE': round(rate * mp, 4)})
        return rows

    # Worst-zone fallback for prefixes not covered by the table
    def _zone_max_rate(zid):
        return max((r for _, r, _ in bands_by_zone.get(zid, [])), default=0)
    all_zone_ids = [z['zone'] for z in zones if 'zone' in z]
    fallback_zone = max(all_zone_ids, key=_zone_max_rate) if all_zone_ids else None

    prefix_to_zone = {}
    for pc_prefix in range(pc_min, pc_max + 1):
        pc_full = pc_prefix * 1000
        matched = next((z['zone'] for z in zones
                        if z.get('pc_from', 0) <= pc_full <= z.get('pc_to', 99999)
                        and 'zone' in z), fallback_zone)
        if matched is not None:
            prefix_to_zone[pc_prefix] = matched

    unique_zones = set(prefix_to_zone.values())
    if len(unique_zones) <= 1:
        zone = next(iter(unique_zones), fallback_zone)
        for each_w, rate, per_kg in bands_by_zone.get(zone, []):
            if each_w > max_ew or per_kg:
                continue
            for mp in range(1, max_p + 1):
                rows.append({**c0, 'POSTCODE': None, 'SERVICE_LEVEL': 'EXPRESS SAVER',
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                             'RATE_BASE': round(rate * mp, 4)})
    else:
        for pc_prefix, zone in prefix_to_zone.items():
            for each_w, rate, per_kg in bands_by_zone.get(zone, []):
                if each_w > max_ew or per_kg:
                    continue
                for mp in range(1, max_p + 1):
                    rows.append({**c0, 'POSTCODE': pc_prefix,
                                 'SERVICE_LEVEL': 'EXPRESS SAVER',
                                 'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                                 'RATE_BASE': round(rate * mp, 4)})
    return rows


def build_rows_postnord(rate_data, country_cfg):
    """
    Handles two PostNord formats:
      Format A — flat_rates dict  {'B2B': 12.2, 'HOME': 12.9, 'PUDO': 10.25}
                 → one row per (service, parcel_count) at max_each_weight
      Format B — weight-tier list {'STANDARD': [{from,to,rate,per_kg},...]}
                 → one row per (weight_band, parcel_count)
    """
    rows   = []
    max_p  = country_cfg['max_parcel_count']
    max_ew = country_cfg['max_each_weight_kg']
    c0     = _common(country_cfg['site_id'], country_cfg['client_id'],
                     'POSTNORD', country_cfg['iso2'])

    flat_rates = rate_data.get('flat_rates', {})
    if flat_rates:
        for svc_name, rate in flat_rates.items():
            for mp in range(1, max_p + 1):
                rows.append({**c0, 'SERVICE_LEVEL': svc_name,
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': max_ew,
                             'RATE_BASE': round(rate * mp, 4)})
    else:
        for each_w, rate, per_kg in collapse_same_rate_tiers(
                rate_data.get('STANDARD', []), max_ew):
            if each_w > max_ew or per_kg:
                continue
            for mp in range(1, max_p + 1):
                rows.append({**c0, 'SERVICE_LEVEL': 'STANDARD',
                             'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                             'RATE_BASE': round(rate * mp, 4)})
    return rows


def build_rows_upsgb(rate_data, country_cfg):
    """UPS GB (UK domestic): STDS (single, per-parcel), STDM (combined weight),
    EXPS (express). Single rate column, no postcode zones. Linehaul applied
    via carrier_defaults."""
    rows   = []
    max_p  = country_cfg['max_parcel_count']
    max_ew = country_cfg['max_each_weight_kg']
    c0     = _common(country_cfg['site_id'], country_cfg['client_id'],
                     'UPSGB', country_cfg['iso2'])
    seen = set()

    # STDS — per-parcel pricing
    for each_w, rate, per_kg in collapse_same_rate_tiers(
            rate_data.get('STDS', []), max_ew):
        if each_w > max_ew or per_kg:
            continue
        for mp in range(1, max_p + 1):
            rb  = rate * mp
            key = (mp, each_w, round(rb, 4))
            if key in seen:
                continue
            seen.add(key)
            rows.append({**c0, 'SERVICE_LEVEL': 'STANDARD',
                         'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                         'RATE_BASE': round(rb, 4)})

    # STDM — combined-weight pricing
    for combined_w, rate, per_kg in collapse_same_rate_tiers(rate_data.get('STDM', [])):
        if per_kg:
            continue
        for mp in range(1, max_p + 1):
            each_w = combined_w / mp
            if each_w not in country_cfg['each_weight_grid'] or each_w > max_ew:
                continue
            key = (mp, each_w, round(rate, 4))
            if key in seen:
                continue
            seen.add(key)
            rows.append({**c0, 'SERVICE_LEVEL': 'STANDARD',
                         'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                         'RATE_BASE': round(rate, 4)})

    # EXPS — express saver
    for each_w, rate, per_kg in collapse_same_rate_tiers(
            rate_data.get('EXPS', []), max_ew):
        if each_w > max_ew or per_kg:
            continue
        for mp in range(1, max_p + 1):
            rows.append({**c0, 'SERVICE_LEVEL': 'EXPRESS SAVER',
                         'MAX_PARCEL': mp, 'EACH_WEIGHT': each_w,
                         'RATE_BASE': round(rate * mp, 4)})
    return rows


CARRIER_BUILDERS = {
    'UPDE':     build_rows_upde,
    'DHL-ROS':  build_rows_dhl,
    'DPD':      build_rows_dpd,
    'UPSNL':    build_rows_upsnl,
    'POSTNORD': build_rows_postnord,
    'UPSGB':    build_rows_upsgb,
}


def build_extended_matrix(parsed, country_cfg):
    all_rows = []
    for carrier in country_cfg['carriers']:
        builder = CARRIER_BUILDERS.get(carrier)
        if builder is None:
            log.warning("No builder for carrier '%s'", carrier)
            continue
        data = parsed.get(carrier, {})
        if not data:
            log.warning("No rate data for '%s', skipping", carrier)
            continue
        carrier_rows = builder(data, country_cfg)
        log.info('  %s: %d rows', carrier, len(carrier_rows))
        all_rows.extend(carrier_rows)
    return pd.DataFrame(all_rows)


# ==============================================================================
# 6. NUMERIC PRE-COMPUTATION
# ==============================================================================

def compute_numeric_totals(df, carrier_defaults=None):
    cd  = carrier_defaults or CARRIER_DEFAULTS
    df  = df.copy()
    if df.empty:
        return df

    df['MAX_WEIGHT'] = df['MAX_PARCEL'] * df['EACH_WEIGHT']

    df['FUEL'] = df.apply(
        lambda r: cd[r['CARRIER_ID']]['fuel_pct'] * r['RATE_BASE'], axis=1
    ).round(4)
    df['MAUT'] = df.apply(
        lambda r: cd[r['CARRIER_ID']]['maut_pct'] * r['RATE_BASE'], axis=1
    ).round(4)
    df['Linehaul UPSDE'] = df.apply(
        lambda r: (cd[r['CARRIER_ID']]['linehaul_per_parcel'] * r['MAX_PARCEL']
                   if cd[r['CARRIER_ID']]['linehaul_per_parcel'] > 0
                   and r['MAX_PARCEL'] is not None
                   and not pd.isna(r['MAX_PARCEL']) else None),
        axis=1,
    ).round(4)

    vdiv = lambda r: cd[r['CARRIER_ID']]['volume_divisor']
    df['MAX_VOLUME']  = df.apply(lambda r: r['MAX_WEIGHT'] / vdiv(r), axis=1)
    df['EACH_VOLUME'] = df.apply(lambda r: r['EACH_WEIGHT'] / vdiv(r), axis=1)

    df['TOTAL_PRICE'] = (
        df['RATE_BASE'] + df['RATE_EXTRA'].fillna(0) + df['FUEL'] + df['MAUT']
        + df['Linehaul UPSDE'].fillna(0)
    ).round(4)
    return df


# ==============================================================================
# 7. EXCEL WRITER
# ==============================================================================

COLUMN_ORDER = [
    'SITE_ID', 'CLIENT_ID', 'CARRIER_ID', 'SERVICE_LEVEL', 'COUNTRYISO2',
    'POSTCODE', 'MIN_WEIGHT', 'MAX_WEIGHT', 'MIN_VOLUME', 'MAX_VOLUME',
    'MIN_PARCEL', 'MAX_PARCEL', 'EACH_WEIGHT', 'EACH_VOLUME',
    'USER_DEF_TYPE_4 (max 1,5m)', 'AWKWARD', 'RATE_BASE', 'RATE_EXTRA',
    'FUEL', 'MAUT', 'Linehaul UPSDE', 'TOTAL_PRICE',
]
COL_LETTER = {name: openpyxl.utils.get_column_letter(i + 1)
              for i, name in enumerate(COLUMN_ORDER)}


def _build_formulas_for_row(row_dict, excel_row, carrier_defaults=None):
    cd  = carrier_defaults or CARRIER_DEFAULTS
    L   = COL_LETTER
    cfg = cd[row_dict['CARRIER_ID']]
    f   = {}
    # Overflow buckets leave MAX_PARCEL / EACH_WEIGHT blank — skip the grid
    # formulas for them so we don't emit "=*" ; their values stay literal (None).
    has_grid = (row_dict.get('MAX_PARCEL') is not None
                and row_dict.get('EACH_WEIGHT') is not None)
    if has_grid:
        f['MAX_WEIGHT']  = f"={L['MAX_PARCEL']}{excel_row}*{L['EACH_WEIGHT']}{excel_row}"
        f['MAX_VOLUME']  = f"={L['MAX_WEIGHT']}{excel_row}/{cfg['volume_divisor']}"
        f['EACH_VOLUME'] = f"={L['EACH_WEIGHT']}{excel_row}/{cfg['volume_divisor']}"
    if cfg.get('fuel_variables_ref'):
        ref = cfg['fuel_variables_ref']
        f['FUEL'] = f"=Variables!${ref[0]}${ref[1:]}*{L['RATE_BASE']}{excel_row}"
    if cfg.get('maut_variables_ref'):
        ref = cfg['maut_variables_ref']
        f['MAUT'] = f"=Variables!${ref[0]}${ref[1:]}*{L['RATE_BASE']}{excel_row}"
    lh = L['Linehaul UPSDE']
    f['TOTAL_PRICE'] = (
        f"={L['RATE_BASE']}{excel_row}+{L['RATE_EXTRA']}{excel_row}"
        f"+{L['FUEL']}{excel_row}+{L['MAUT']}{excel_row}"
        f'+{lh}{excel_row}'
    )
    return f


def write_matrix_excel(df, output_path, country_cfg,
                       carrier_defaults=None, variables_layout=None):
    vl  = variables_layout or VARIABLES_LAYOUT
    wb  = Workbook()
    ws  = wb.active
    ws.title = f"{country_cfg['iso2']} Matrix"
    for ci, col in enumerate(COLUMN_ORDER, 1):
        ws.cell(1, ci, col)
    df_sorted = df.sort_values('TOTAL_PRICE', kind='stable').reset_index(drop=True)
    bucket_fill = PatternFill('solid', fgColor='FFF2CC')   # soft amber = catch-all bucket
    for ri, row_dict in enumerate(df_sorted.to_dict('records'), start=2):
        formulas = _build_formulas_for_row(row_dict, ri, carrier_defaults)
        is_bucket = bool(row_dict.get('_is_bucket'))
        for ci, col in enumerate(COLUMN_ORDER, 1):
            if col in formulas:
                cell = ws.cell(ri, ci, formulas[col])
            else:
                val = row_dict.get(col)
                cell = ws.cell(ri, ci, None if pd.isna(val) else val)
            if is_bucket:
                cell.fill = bucket_fill
    vs = wb.create_sheet('Variables')
    for ri, (name, val) in enumerate(vl, 1):
        vs.cell(ri, 1, name)
        vs.cell(ri, 2, val)
    wb.save(output_path)
    log.info('wrote %s (%d rows)', output_path, len(df_sorted))


# ==============================================================================
# 8. FIRST-PASS OPTIMIZER (per carrier/service)
# ==============================================================================

def optimize_matrix(df):
    df = df.copy()
    df['_origrow'] = df.index
    keep_mask = pd.Series(True, index=df.index)
    df_filled = df.copy()
    df_filled['POSTCODE_KEY'] = (df_filled['POSTCODE'].fillna(-1)
                                 .infer_objects(copy=False))
    for _, grp in df_filled.groupby(['CARRIER_ID', 'SERVICE_LEVEL', 'POSTCODE_KEY']):
        grp = grp.sort_values('TOTAL_PRICE', kind='stable').reset_index()
        for i in range(len(grp)):
            r       = grp.iloc[i]
            earlier = grp.iloc[:i]
            dom = ((earlier['MAX_WEIGHT']  >= r['MAX_WEIGHT']) &
                   (earlier['MAX_PARCEL']  >= r['MAX_PARCEL']) &
                   (earlier['EACH_WEIGHT'] >= r['EACH_WEIGHT']))
            if dom.any():
                keep_mask.loc[r['_origrow']] = False
    log.info('first-pass: removed %d dominated rows', (~keep_mask).sum())
    return df[keep_mask].drop(columns=['_origrow']).reset_index(drop=True)


# ==============================================================================
# 9. GLOBAL OPTIMIZER (cross-carrier)
# ==============================================================================

_RELATIVE_REF = re.compile(r'(\$?[A-Z]+)(\$?)(\d+)')


def _update_formula(formula, old_row, new_row):
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula
    def repl(m):
        col, dollar, row = m.group(1), m.group(2), m.group(3)
        if dollar == '$':
            return m.group(0)
        return f'{col}{new_row}' if int(row) == old_row else m.group(0)
    return _RELATIVE_REF.sub(repl, formula)


def _write_filtered_excel(input_path, output_path, keep_indices):
    shutil.copy(input_path, output_path)
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    ncols = ws.max_column
    keep_excel = {idx + 2 for idx in keep_indices}
    kept = [(r, [ws.cell(r, c).value for c in range(1, ncols + 1)])
            for r in range(2, ws.max_row + 1) if r in keep_excel]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for new_row, (old_row, vals) in enumerate(kept, start=2):
        for c, val in enumerate(vals, start=1):
            if isinstance(val, str) and val.startswith('='):
                val = _update_formula(val, old_row, new_row)
            ws.cell(new_row, c, val)
    wb.save(output_path)
    return len(kept)


def _ensure_numeric(df, input_path):
    df = df.copy()
    if df['MAX_WEIGHT'].isna().any():
        df['MAX_WEIGHT'] = df['MAX_PARCEL'] * df['EACH_WEIGHT']
    variables = {}
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        if 'Variables' in wb.sheetnames:
            vs = wb['Variables']
            for r in range(1, vs.max_row + 1):
                name = vs.cell(r, 1).value
                val  = vs.cell(r, 2).value
                if isinstance(name, str) and isinstance(val, (int, float)):
                    variables[name.strip()] = float(val)
    except Exception as e:
        log.warning('Could not read Variables sheet: %s', e)
    fuel_map = {'UPDE': 'FUEL UPSDE', 'DHL-ROS': 'FUEL DHL',
                'DPD': 'FUEL DPD', 'UPSNL': 'FUEL UPSNL',
                'POSTNORD': 'FUEL POSTNORD'}
    maut_map = {'DPD': 'MAUT DPD', 'DHL-ROS': 'MAUT DHL'}
    pct = lambda carrier, mapping: variables.get(mapping.get(carrier, ''), 0.0)
    if df['FUEL'].isna().any():
        df['FUEL'] = df.apply(
            lambda r: pct(r['CARRIER_ID'], fuel_map) * (r['RATE_BASE'] or 0), axis=1)
    if df['MAUT'].isna().any():
        df['MAUT'] = df.apply(
            lambda r: pct(r['CARRIER_ID'], maut_map) * (r['RATE_BASE'] or 0), axis=1)
    if df['TOTAL_PRICE'].isna().any():
        df['TOTAL_PRICE'] = (df['RATE_BASE'].fillna(0) + df['RATE_EXTRA'].fillna(0)
                             + df['FUEL'].fillna(0) + df['MAUT'].fillna(0)
                             + df['Linehaul UPSDE'].fillna(0))
    return df


def optimize_globally(input_path, output_path):
    df = pd.read_excel(input_path, sheet_name=0)
    df = _ensure_numeric(df, input_path)
    df_s = df.sort_values('TOTAL_PRICE', kind='stable').reset_index()
    df_s = df_s.rename(columns={'index': '_orig'})
    w  = df_s['MAX_WEIGHT'].values.astype(float)
    p  = df_s['MAX_PARCEL'].values.astype(float)
    e  = df_s['EACH_WEIGHT'].values.astype(float)
    pc = df_s['POSTCODE'].values.astype(float)
    pc_nan = np.isnan(pc)
    orig   = df_s['_orig'].values
    dominated = set()
    for i in range(1, len(df_s)):
        pc_compat = pc_nan[:i] if pc_nan[i] else (pc_nan[:i] | (pc[:i] == pc[i]))
        if (pc_compat & (w[:i] >= w[i]) & (p[:i] >= p[i]) & (e[:i] >= e[i])).any():
            dominated.add(int(orig[i]))
    keep = set(df.index) - dominated
    n = _write_filtered_excel(input_path, output_path, keep)
    log.info('global optimizer: %d → %d rows (%d removed)', len(df), n, len(dominated))
    return {'input_rows': len(df), 'removed': len(dominated),
            'output_rows': n, 'output_path': str(output_path)}


# ==============================================================================
# 9b. EXCEPTIONS / BUCKETS  (add-on)
# ==============================================================================
#
# CargoWrite matches an order by scanning the matrix top-to-bottom (cheapest
# first) and taking the first row whose constraints all fit. A constraint such
# as USER_DEF_TYPE_4 (a max dimension) means "this row only matches parcels at
# or under this size". Orders that exceed it must still match *something*, so
# every constrained row needs a cheaper-to-build "bucket" twin lower down that
# drops the limit, flags the row for oversight, and adds a surcharge.
#
# An exception rule is a plain dict (general — not size-specific):
#   {
#     'enabled':        True,
#     'label':          'Oversize (max 1.5m)',
#     'carriers':       ['UPDE'],      # scope; [] = all carriers
#     'countries':      [],            # scope; [] = all countries
#     'constraint_col': 'USER_DEF_TYPE_4 (max 1,5m)',
#     'normal_value':   1.5,           # stamped on the cheap base rows
#     'bucket_value':   None,          # value on the bucket twin (None = catch-all)
#     'flag_col':       'AWKWARD',     # column flagged on the bucket twin
#     'flag_value':     'y',
#     'surcharge':      6.0,           # euros
#     'surcharge_mode': 'per_parcel',  # 'per_parcel' (× MAX_PARCEL) or 'flat'
#   }
#
# FUTURE BUCKET IDEAS (designed for, not yet implemented):
#   • parcel-count overflow  — a row with MAX_PARCEL blank to catch >max parcels
#   • weight overflow        — a row with MAX_WEIGHT blank for over-grid weights
#   • postcode catch-all     — a blank-POSTCODE row when zone rows miss a prefix
#   apply_exceptions already leaves room for these: add a rule whose
#   constraint_col is MAX_PARCEL / MAX_WEIGHT / POSTCODE and bucket_value=None.


def _recompute_total(df):
    df['TOTAL_PRICE'] = (
        df['RATE_BASE'].fillna(0) + df['RATE_EXTRA'].fillna(0)
        + df['FUEL'].fillna(0) + df['MAUT'].fillna(0)
        + df['Linehaul UPSDE'].fillna(0)
    ).round(4)
    return df


def apply_exceptions(df, rules):
    """Stamp normal limits on in-scope base rows and append bucket twins.

    Returns a new DataFrame with an extra boolean column '_is_bucket'
    (write_matrix_excel uses it to colour the bucket rows). The base matrix is
    expected to already carry numeric FUEL / MAUT / Linehaul columns.
    """
    if not rules:
        return df

    df = df.copy().reset_index(drop=True)
    if '_is_bucket' not in df.columns:
        df['_is_bucket'] = False

    new_buckets = []
    for rule in rules:
        if not rule.get('enabled', True):
            continue
        ccol = rule['constraint_col']
        if ccol not in df.columns:
            log.warning("exception rule skipped — column '%s' not in matrix", ccol)
            continue

        scope = (~df['_is_bucket'])
        if rule.get('carriers'):
            scope &= df['CARRIER_ID'].isin(rule['carriers'])
        if rule.get('countries'):
            scope &= df['COUNTRYISO2'].isin(rule['countries'])
        if not scope.any():
            continue

        # 1) stamp the normal limit onto the cheap base rows
        df.loc[scope, ccol] = rule['normal_value']

        # 2) build the bucket twins (limit removed, flagged, surcharged)
        twins = df[scope].copy()
        twins[ccol] = rule.get('bucket_value', None)
        if rule.get('flag_col'):
            twins[rule['flag_col']] = rule.get('flag_value', 'y')
        sur = float(rule.get('surcharge', 0) or 0)
        if rule.get('surcharge_mode', 'per_parcel') == 'per_parcel':
            twins['RATE_EXTRA'] = twins['RATE_EXTRA'].fillna(0) + sur * twins['MAX_PARCEL']
        else:
            twins['RATE_EXTRA'] = twins['RATE_EXTRA'].fillna(0) + sur
        twins['_is_bucket'] = True
        new_buckets.append(twins)

    if new_buckets:
        df = pd.concat([df] + new_buckets, ignore_index=True)
        df = _recompute_total(df)
    return df


def add_overflow_buckets(df, rules, carrier_defaults=None, country_cfg=None):
    """Append combined parcel+weight overflow buckets (matches the corrected
    example's structure). For each in-scope (carrier, service, country) and each
    parcel count n in 1..grid_max_parcels, add a catch-all row:

        MIN_PARCEL = n,            MAX_PARCEL = blank   (catches >= n parcels)
        MIN_WEIGHT = n*grid_max,   MAX_WEIGHT = blank   (catches over-grid weight)
        EACH_WEIGHT = blank
        RATE_BASE  = overflow_rate * n       (overflow_rate is manager-supplied)
        RATE_EXTRA = surcharge * n
        flag       = AWKWARD = 'Y'

    The overflow_rate is NOT inferred from the grid (the example used a hand-set
    heavy/per-kg rate); it must be provided in the rule. Rows are flagged so ops
    can review every overflow shipment.
    """
    cd = carrier_defaults or CARRIER_DEFAULTS
    if not rules:
        return df
    df = df.copy().reset_index(drop=True)
    if '_is_bucket' not in df.columns:
        df['_is_bucket'] = False

    new = []
    for rule in rules:
        if not rule.get('enabled', True):
            continue
        try:
            rate = float(rule['overflow_rate'])
        except (TypeError, ValueError, KeyError):
            log.warning('overflow rule skipped — no valid overflow_rate'); continue
        sur      = float(rule.get('surcharge', 0) or 0)
        flag_col = rule.get('flag_col', 'AWKWARD')
        flag_val = rule.get('flag_value', 'Y')

        base = df[~df['_is_bucket']]
        if rule.get('carriers'):
            base = base[base['CARRIER_ID'].isin(rule['carriers'])]
        if rule.get('countries'):
            base = base[base['COUNTRYISO2'].isin(rule['countries'])]
        if base.empty:
            continue

        for (carrier, svc, country), grp in base.groupby(
                ['CARRIER_ID', 'SERVICE_LEVEL', 'COUNTRYISO2']):
            # Grid ceilings come from the country config (the TRUE limits), not
            # the post-optimization frame where weight bands have collapsed.
            if country_cfg:
                max_p    = int(country_cfg['max_parcel_count'])
                grid_max = float(country_cfg['max_each_weight_kg'])
            else:
                max_p    = int(grp['MAX_PARCEL'].max())
                grid_max = float(grp['EACH_WEIGHT'].max())
            site, client = grp.iloc[0]['SITE_ID'], grp.iloc[0]['CLIENT_ID']
            for n in range(1, max_p + 1):
                row = _common(site, client, carrier, country)
                rb  = round(rate * n, 4)
                fuel = round(cd[carrier]['fuel_pct'] * rb, 4)
                maut = round(cd[carrier]['maut_pct'] * rb, 4)
                lh_pp = cd[carrier]['linehaul_per_parcel']
                lh   = round(lh_pp * n, 4) if lh_pp > 0 else None
                row.update({
                    'SERVICE_LEVEL': svc,
                    'MIN_PARCEL': n,       'MAX_PARCEL': None,
                    'MIN_WEIGHT': round(n * grid_max, 4), 'MAX_WEIGHT': None,
                    'EACH_WEIGHT': None,   'MAX_VOLUME': None, 'EACH_VOLUME': None,
                    'RATE_BASE': rb,       'RATE_EXTRA': round(sur * n, 4),
                    'FUEL': fuel,          'MAUT': maut, 'Linehaul UPSDE': lh,
                    flag_col: flag_val,    '_is_bucket': True,
                })
                row['TOTAL_PRICE'] = round(rb + sur * n + fuel + maut + (lh or 0), 4)
                new.append(row)

    if new:
        df = pd.concat([df, pd.DataFrame(new)], ignore_index=True)
    return df


def add_postcode_catchall(df, rules, carrier_defaults=None):
    """For zoned carriers (rows carrying a specific POSTCODE prefix), append a
    POSTCODE=blank fallback at the worst (most expensive) zone's rate, flagged,
    so a prefix not present in any zone still matches something."""
    if not rules:
        return df
    df = df.copy().reset_index(drop=True)
    if '_is_bucket' not in df.columns:
        df['_is_bucket'] = False

    new = []
    for rule in rules:
        if not rule.get('enabled', True):
            continue
        sur      = float(rule.get('surcharge', 0) or 0)
        flag_col = rule.get('flag_col', 'AWKWARD')
        flag_val = rule.get('flag_value', 'Y')

        base = df[(~df['_is_bucket']) & (df['POSTCODE'].notna())]
        if rule.get('carriers'):
            base = base[base['CARRIER_ID'].isin(rule['carriers'])]
        if rule.get('countries'):
            base = base[base['COUNTRYISO2'].isin(rule['countries'])]
        if base.empty:
            continue

        # worst-case row per (carrier, service, country, parcels, each-weight)
        keys = ['CARRIER_ID', 'SERVICE_LEVEL', 'COUNTRYISO2', 'MAX_PARCEL', 'EACH_WEIGHT']
        worst = base.loc[base.groupby(keys)['TOTAL_PRICE'].idxmax()].copy()
        worst['POSTCODE']   = None
        worst['RATE_EXTRA'] = worst['RATE_EXTRA'].fillna(0) + sur
        worst[flag_col]     = flag_val
        worst['_is_bucket'] = True
        new.append(worst)

    if new:
        df = pd.concat([df] + new, ignore_index=True)
        df = _recompute_total(df)
    return df


def optimize_globally_df(df):
    """Cross-carrier dominance on a DataFrame; returns the kept rows.
    Same logic as optimize_globally but stays in pandas (no Excel round-trip),
    so the exception/bucket step can run on the result before writing."""
    df = df.reset_index(drop=True)
    if df.empty:
        return df
    df_s = df.sort_values('TOTAL_PRICE', kind='stable').reset_index()
    df_s = df_s.rename(columns={'index': '_orig'})
    w  = df_s['MAX_WEIGHT'].values.astype(float)
    p  = df_s['MAX_PARCEL'].values.astype(float)
    e  = df_s['EACH_WEIGHT'].values.astype(float)
    pc = df_s['POSTCODE'].values.astype(float)
    pc_nan = np.isnan(pc)
    orig = df_s['_orig'].values
    dominated = set()
    for i in range(1, len(df_s)):
        pc_compat = pc_nan[:i] if pc_nan[i] else (pc_nan[:i] | (pc[:i] == pc[i]))
        if (pc_compat & (w[:i] >= w[i]) & (p[:i] >= p[i]) & (e[:i] >= e[i])).any():
            dominated.add(int(orig[i]))
    keep = [i for i in range(len(df)) if i not in dominated]
    return df.iloc[keep].reset_index(drop=True)


# ==============================================================================
# 10. ORCHESTRATOR
# ==============================================================================

def run_pipeline(input_path, country, output_dir='.',
                 country_cfg=None, carrier_defaults=None, variables_layout=None,
                 exceptions=None, overflow_rules=None, postcode_rules=None):
    """
    Full pipeline for one country.

    Parameters
    ----------
    input_path       : path to the rate-card Excel
    country          : ISO-2 code, e.g. 'DE'
    output_dir       : directory where output files are written
    country_cfg      : dict — overrides COUNTRY_CONFIG[country] entirely
    carrier_defaults : dict — overrides module-level CARRIER_DEFAULTS
    variables_layout : list — overrides module-level VARIABLES_LAYOUT

    Returns
    -------
    dict with keys: extended, optimized, minimal,
                    rows_extended, rows_optimized, rows_minimal
    """
    country = country.upper()
    cfg = country_cfg or COUNTRY_CONFIG.get(country)
    if cfg is None:
        raise ValueError(f"No configuration for country '{country}'.")
    cd  = carrier_defaults or CARRIER_DEFAULTS
    vl  = variables_layout or VARIABLES_LAYOUT

    log.info('=== Pipeline for %s ===', country)
    parsed = parse_rate_cards(input_path)
    return run_pipeline_from_parsed(parsed, country, output_dir, cfg, cd, vl,
                                    exceptions, overflow_rules, postcode_rules)


def run_pipeline_from_parsed(parsed, country, output_dir, cfg,
                             carrier_defaults=None, variables_layout=None,
                             exceptions=None, overflow_rules=None,
                             postcode_rules=None):
    """Build/optimize/write from an already-parsed rate dict.
    Used by the master-file path so the (expensive) parse happens only once.

    If `exceptions` (a list of rule dicts) is given, each output also gets the
    stamped limits + catch-all bucket rows (coloured), and the minimal stage
    runs fully in pandas so buckets are added to the *surviving* rows only —
    keeping the list as short as possible.
    """
    cd = carrier_defaults or CARRIER_DEFAULTS
    vl = variables_layout or VARIABLES_LAYOUT
    country = country.upper()

    df = build_extended_matrix(parsed, cfg)
    log.info('raw rows: %d', len(df))
    if df.empty:
        raise ValueError(f"No rows built for {country} — no matching rate data "
                         f"for carriers {cfg['carriers']}.")

    df = compute_numeric_totals(df, cd)
    out = Path(output_dir)
    ext_path = out / f'{country}_Matrix_extended.xlsx'
    opt_path = out / f'{country}_Matrix_optimized.xlsx'
    min_path = out / f'{country}_Matrix_minimal.xlsx'

    df_opt = optimize_matrix(df)

    any_buckets = bool(exceptions or overflow_rules or postcode_rules)
    if any_buckets:
        # buckets added to surviving rows of each stage, written fresh (pandas flow).
        # Order: overflow + postcode catch-alls first, then size buckets last so the
        # size rule (which skips existing buckets) only stamps the normal grid rows.
        df_min = optimize_globally_df(df_opt)

        def _decorate(d):
            d = add_overflow_buckets(d, overflow_rules, cd, cfg)
            d = add_postcode_catchall(d, postcode_rules, cd)
            d = apply_exceptions(d, exceptions)
            return d

        ext_b, opt_b, min_b = _decorate(df), _decorate(df_opt), _decorate(df_min)
        write_matrix_excel(ext_b, ext_path, cfg, cd, vl)
        write_matrix_excel(opt_b, opt_path, cfg, cd, vl)
        write_matrix_excel(min_b, min_path, cfg, cd, vl)
        rows_ext, rows_opt, rows_min = len(ext_b), len(opt_b), len(min_b)
    else:
        write_matrix_excel(df, ext_path, cfg, cd, vl)
        write_matrix_excel(df_opt, opt_path, cfg, cd, vl)
        stats = optimize_globally(opt_path, min_path)
        rows_ext, rows_opt, rows_min = len(df), len(df_opt), stats['output_rows']

    log.info('=== Done: %d ext / %d opt / %d min ===', rows_ext, rows_opt, rows_min)
    return {
        'extended':      str(ext_path),
        'optimized':     str(opt_path),
        'minimal':       str(min_path),
        'rows_extended': rows_ext,
        'rows_optimized': rows_opt,
        'rows_minimal':  rows_min,
    }
